import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
try:
    from exporters.matrix_gen import MatrixGen
except ModuleNotFoundError:
    from ..exporters.matrix_gen import MatrixGen


class BadTemplateException(Exception):
    pass

class ExcelTemplates:

    def __init__(self, source='taxii', local=None, domain='enterprise'):
        """
            Initialization - Creates a ExcelTemplate object

            :param source: Source to use when compiling the matrix
            :param local: Optional path to local stix data, required if source == "local"
            :param domain: The domain to utilize
        """
        muse = domain
        if muse.startswith('mitre-'):
            muse = domain[6:]
        if muse.endswith('-attack'):
            muse = domain[:-7]
        if muse in ['enterprise', 'mobile']:
            self.mode = muse
            self.h = MatrixGen(source=source, local=local)
            self.codex = self.h.get_matrix(muse)
        else:
            raise BadTemplateException

    def _build_raw(self, showName=True, showID=False, sort=0, scores=[], subtechs=[], exclude=[]):
        """
            INTERNAL - builds a raw, not-yet-marked-up excel document based on the specifications

            :param showName: Whether or not to display names for each entry
            :param showID: Whether or not to display Technique IDs for each entry
            :param sort: The sort mode to use
            :param subtechs: List of all visible subtechniques
            :param exclude: List of of techniques to exclude from the matrix
            :return: a openpyxl workbook object containing the raw matrix
        """
        self.codex = self.h._adjust_ordering(self.codex, sort, scores)
        template, joins = self.h._construct_panop(self.codex, subtechs, exclude)
        self.template = template
        wb = openpyxl.Workbook()

        sheet = wb.active

        header_template_f = Font(name='Calibri', bold=True)
        header_template_a = Alignment(horizontal='center', vertical='bottom')
        header_template_b = Border(bottom=Side(border_style='thin'))
        header_template_c = PatternFill(patternType='solid', start_color='DDDDDD', end_color='DDDDDD')

        for entry in template:
            c = sheet.cell(row=entry[0], column=entry[1])
            write_val = ''
            if showName and showID:
                write_val = self.h._get_ID(self.codex, template[entry]) + ': ' + template[entry]
            elif showName:
                write_val = template[entry]
            elif showID:
                write_val = self.h._get_ID(self.codex, template[entry])
            c.value = write_val
            if entry[0] == 1:
                c.font = header_template_f
                c.alignment = header_template_a
                c.border = header_template_b
                c.fill = header_template_c

        ## patch widths
        dims = {}
        sheet_handle = wb.active
        for row in sheet_handle:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            sheet_handle.column_dimensions[col].width = value

        merge_border_thickness = 'thin'
        merge_template_l = Border(bottom=Side(border_style=merge_border_thickness),
                                  left=Side(border_style=merge_border_thickness),
                                  top=Side(border_style=merge_border_thickness))
        merge_template_r = Border(bottom=Side(border_style=merge_border_thickness),
                                  right=Side(border_style=merge_border_thickness),
                                  top=Side(border_style=merge_border_thickness))

        for marker in joins:
            sheet_handle.merge_cells(start_row=marker[0], start_column=marker[1], end_row=marker[0] + marker[2] - 1,
                                     end_column=marker[1])
            for block in range(marker[0], marker[0] + marker[2]):
                sheet_handle[block][marker[1]].border = merge_template_r
                sheet_handle[block][marker[1] - 1].border = merge_template_l
            sheet_handle.merge_cells(start_row=1, start_column=marker[1], end_row=1, end_column=marker[1] + 1)
            adjust = sheet_handle.cell(row=marker[0], column=marker[1])
            adjust.alignment = Alignment(vertical='top')

        return wb

    def export(self, showName, showID, sort=0, scores=[], subtechs=[], exclude=[]):
        """
            Export a raw customized excel template


            :param showName: Whether or not to display names for each entry
            :param showID: Whether or not to display Technique IDs for each entry
            :param sort: The sort mode to utilize
            :param subtechs: List of all visible subtechniques
            :param exclude: List of of techniques to exclude from the matrix
            return: a openpyxl workbook object containing the raw matrix
        """
        return self._build_raw(showName, showID, sort, scores, subtechs, exclude)

    def retrieve_coords(self, techniqueID, tactic=None):
        """
            Locate the openpyxl coordinates of the provided technique for the currently loaded matrix

            :param techniqueID: The ID of the technique to locate
            :param tactic: Optional parameter to isolate the technique to a specific tactic
            :return: A tuple representing the (row, column) of the target element in the workbook
        """
        listing = []
        match = self.h._get_name(self.codex, techniqueID)
        for entry in self.template:
            if self.template[entry] == match:
                if tactic is not None:
                    try:
                        if self.template[(1, entry[1])] != self.h.convert(tactic):
                            continue
                    except KeyError:
                        # account for subtechniques when scanning
                        if self.template[(1, entry[1] - 1)] != self.h.convert(tactic):
                            continue
                listing.append(entry)
        if listing == []:
            if '.' in techniqueID:
                parent = self.retrieve_coords(techniqueID.split('.')[0], tactic)
                if parent != []:
                    return 'HIDDEN'
        return listing
