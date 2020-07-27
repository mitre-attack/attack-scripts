from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font
import colorsys

try:
    from core import Layer
    from exporters import ExcelTemplates
except ModuleNotFoundError:
    from ..core import Layer
    from ..exporters import ExcelTemplates

class ToExcel:
    def __init__(self, domain='enterprise', source='taxii', local=None):
        """
            Sets up exporting system, builds underlying matrix

            :param source: Source to generate the matrix from, one of (taxii or local)
            :param local: Optional path to local stix data, required in source is local

        """
        self.domain = domain
        self.raw_handle = ExcelTemplates(domain=domain, source=source, local=local)

    def to_xlsx(self, layer, filepath="layer.xlsx"):
        """
            Exports a layer file to the excel format as the file specified

            :param layer: A layer to initialize the instance with
            :param filepath: The location to write the excel file to
        """

        if not isinstance(layer, Layer):
            raise TypeError
            
        if self.domain not in layer.layer.domain:
            raise ValueError(f"layer domain ({layer.layer.domain}) does not match exporter domain ({self.domain})")

        included_subs = []
        if layer.layer.techniques:
            for entry in layer.layer.techniques:
                if entry.showSubtechniques:
                    if entry.tactic:
                        included_subs.append((entry.techniqueID, entry.tactic))
                    else:
                        included_subs.append((entry.techniqueID, False))

        excluded = []
        if layer.layer.hideDisabled:
            for entry in layer.layer.techniques:
                if entry.enabled == False:
                    if entry.tactic:
                        excluded.append((entry.techniqueID, entry.tactic))
                    else:
                        excluded.append((entry.techniqueID, False))
        scores = []
        if layer.layer.techniques:
            for entry in layer.layer.techniques:
                if entry.score:
                    if entry.tactic:
                        scores.append((entry.techniqueID, entry.tactic, entry.score))
                    else:
                        scores.append((entry.techniqueID, False, entry.score))
        sName = True
        sID = False
        sort = 0
        if layer.layer.layout:
            sName = layer.layer.layout.showName
            sID = layer.layer.layout.showID
        if layer.layer.sorting:
            sort = layer.layer.sorting
        raw_template = self.raw_handle.export(showName=sName, showID=sID, sort=sort, scores=scores,
                                              subtechs=included_subs, exclude=excluded)
        sheet_obj = raw_template.active
        sheet_obj.title = layer.layer.name
        for tech in layer.layer.techniques:
            p_tactic = None
            if tech.tactic:
                p_tactic = tech.tactic
            coords = self.raw_handle.retrieve_coords(tech.techniqueID, p_tactic)
            if coords == [] or coords == 'HIDDEN':
                tac = p_tactic
                if tac == None:
                    tac = "(none)"
                if coords == []:
                    print('WARNING! Technique/Tactic ' + tech.techniqueID + '/' + tac +
                          ' does not appear to exist in the loaded matrix. Skipping...')
                else:
                    parents = [x for x in layer.layer.techniques if x.techniqueID == tech.techniqueID.split('.')[0]]
                    if tech.tactic:
                        parents = [x for x in parents if x.tactic == tech.tactic]
                    if all([True if not x.showSubtechniques else False for x in parents]):
                        print('NOTE! Technique/Tactic ' + tech.techniqueID + '/' + tac + ' does not appear '
                              'to be visible in the matrix. Its parent appears to be hiding it.')
                    else:
                        print('WARNING! Technique/Tactic ' + tech.techniqueID + '/' + tac + ' seems malformed. '
                              'Skipping...')
                    continue
            for location in coords:
                cell = sheet_obj.cell(row=location[0],column=location[1])
                if tech.comment:
                    cell.comment = Comment(tech.comment, 'ATT&CK Scripts Exporter')

                if tech.enabled == False:
                    if layer.layer.hideDisabled:
                        pass
                    else:
                        grayed_out = Font(color='909090')
                        cell.font = grayed_out
                        continue
                if tech.color:
                    c_color = PatternFill(fill_type='solid', start_color=tech.color.upper()[1:])
                    cell.fill = c_color
                    continue
                if tech.score:
                    if layer.layer.gradient:
                        comp_color = layer.layer.gradient.compute_color(tech.score)
                        c_color = PatternFill(fill_type='solid', start_color=comp_color.upper()[1:])
                        cell.fill = c_color
                        RGB = tuple(int(comp_color.upper()[1:][i:i+2], 16) for i in (0, 2, 4))
                        hls = colorsys.rgb_to_hls(RGB[0], RGB[1], RGB[2])
                        if hls[1] < 127.5:
                            white = Font(color='FFFFFF')
                            cell.font = white
        raw_template.save(filepath)
